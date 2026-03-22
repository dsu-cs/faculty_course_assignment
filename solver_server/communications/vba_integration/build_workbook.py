"""
Excel Workbook Builder
Automatically creates the complete Faculty Course Assignment workbook
Author: Muhammad Bhutta
"""

import openpyxl
from openpyxl import Workbook
import os

def create_workbook():
    """Create complete Excel workbook with sample data"""
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets
    preferences_sheet = wb.create_sheet("Preferences")
    timeblocks_sheet = wb.create_sheet("TimeBlocks")
    workload_sheet = wb.create_sheet("Workload")
    results_sheet = wb.create_sheet("Results")
    
    # Load and populate Preferences
    with open('sample_data/preferences_sample.csv', 'r') as f:
        for row_idx, line in enumerate(f, 1):
            values = line.strip().split(',')
            for col_idx, value in enumerate(values, 1):
                preferences_sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Load and populate TimeBlocks
    with open('sample_data/time_blocks_sample.csv', 'r') as f:
        for row_idx, line in enumerate(f, 1):
            values = line.strip().split(',')
            for col_idx, value in enumerate(values, 1):
                timeblocks_sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Load and populate Workload
    with open('sample_data/workload_sample.csv', 'r') as f:
        for row_idx, line in enumerate(f, 1):
            values = line.strip().split(',')
            for col_idx, value in enumerate(values, 1):
                workload_sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Add headers to Results sheet
    results_sheet.cell(1, 1, "Results will appear here after running solver")
    
    # Save workbook
    output_path = 'FacultyCourseAssignment.xlsx'
    wb.save(output_path)
    print(f"Workbook created: {output_path}")
    print("\nNext steps:")
    print("1. Open the workbook in Excel")
    print("2. Save as .xlsm (macro-enabled)")
    print("3. Import the three .bas files from VBA Editor")
    print("4. Add button assigned to MainController.RunSolver")

if __name__ == '__main__':
    os.chdir(os.path.dirname(__file__))
    create_workbook()
