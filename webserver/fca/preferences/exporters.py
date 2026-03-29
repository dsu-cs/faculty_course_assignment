from __future__ import annotations

import csv
import importlib.util
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from django.conf import settings

from fca.preferences.models import FacultyPreferenceSubmission
from fca.preferences.services import RAW_SECTION_HEADERS
from fca.preferences.services import load_raw_sections_tab_data
from fca.preferences.services import load_sections_tab_data


REFERENCE_WORKBOOK_DIR_CANDIDATES = (
    Path("/reference-workbook"),
    Path(settings.BASE_DIR) / "Reference Workbook",
    Path(settings.BASE_DIR).parent / "Reference Workbook",
)
REFERENCE_WORKBOOK_FILE_CANDIDATES = ("workbook.xlsm", "workbook.xlsx")
BIM_SCRAPER_PATH = Path(settings.BASE_DIR) / "BIM Scraper" / "bim_scraper.py"
EXPORT_DIR = Path(settings.MEDIA_ROOT) / "dean_downloads"
PREFERENCE_RANK = {"X": -1, "0": 0, "1": 1, "2": 2, "3": 3}


@dataclass
class DeanDownloadArtifacts:
    sections_csv_path: Path
    preferences_csv_path: Path
    workbook_path: Path


def _normalize_preference_for_csv(value: str) -> str:
    return "x" if value.upper() == "X" else value


def _get_reference_workbook_dir() -> Path:
    for candidate in REFERENCE_WORKBOOK_DIR_CANDIDATES:
        if candidate.exists():
            return candidate
    searched_paths = ", ".join(str(path) for path in REFERENCE_WORKBOOK_DIR_CANDIDATES)
    msg = f"Reference Workbook folder not found. Looked in: {searched_paths}"
    raise FileNotFoundError(msg)


def _get_reference_workbook_path() -> Path:
    workbook_dir = _get_reference_workbook_dir()
    for filename in REFERENCE_WORKBOOK_FILE_CANDIDATES:
        candidate = workbook_dir / filename
        if candidate.exists():
            return candidate
    searched_paths = ", ".join(str(workbook_dir / filename) for filename in REFERENCE_WORKBOOK_FILE_CANDIDATES)
    msg = f"Reference workbook not found. Looked in: {searched_paths}"
    raise FileNotFoundError(msg)


def _load_module(module_name: str, module_path: Path):
    spec = importlib.util.spec_from_file_location(module_name, module_path)
    if spec is None or spec.loader is None:
        msg = f"Unable to load module from {module_path}"
        raise ImportError(msg)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def get_available_bim_terms() -> list[str]:
    scraper_module = _load_module("bim_scraper", BIM_SCRAPER_PATH)
    return list(scraper_module.fetch_available_terms())


def get_default_bim_term(available_terms: list[str]) -> str | None:
    if not available_terms:
        return None
    scraper_module = _load_module("bim_scraper", BIM_SCRAPER_PATH)
    chosen_term, _warning = scraper_module.select_term(None, available_terms)
    return chosen_term


def refresh_bim_sections_csv(output_path: Path, term: str | None = None) -> Path:
    scraper_module = _load_module("bim_scraper", BIM_SCRAPER_PATH)
    return Path(scraper_module.run(term=term, output_path=output_path))


def _latest_submissions_by_faculty() -> dict[str, FacultyPreferenceSubmission]:
    submissions = (
        FacultyPreferenceSubmission.objects.prefetch_related("course_preferences")
        .order_by("faculty_identifier", "-submitted_at", "-id")
    )

    latest: dict[str, FacultyPreferenceSubmission] = {}
    for submission in submissions:
        faculty_key = submission.faculty_identifier.strip() or "anonymous"
        if faculty_key not in latest:
            latest[faculty_key] = submission
    return latest


def _collapse_submission_preferences(submission: FacultyPreferenceSubmission) -> dict[tuple[str, str], str]:
    grouped_preferences: dict[tuple[str, str], str] = {}

    for course_preference in submission.course_preferences.all():
        key = (course_preference.prefix, course_preference.course_number)
        existing_value = grouped_preferences.get(key, "X")
        if PREFERENCE_RANK[course_preference.preference] >= PREFERENCE_RANK[existing_value]:
            grouped_preferences[key] = course_preference.preference

    return grouped_preferences


def build_preferences_csv(output_path: Path, sections_csv_path: Path | None = None) -> Path:
    sections = load_sections_tab_data(sections_csv_path) if sections_csv_path else load_sections_tab_data()
    latest_submissions = _latest_submissions_by_faculty()
    faculty_names = list(latest_submissions.keys())

    if not faculty_names:
        msg = "No faculty preference submissions were found in Postgres."
        raise ValueError(msg)

    preference_maps = {
        faculty_name: _collapse_submission_preferences(submission)
        for faculty_name, submission in latest_submissions.items()
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(["section_id", *faculty_names])

        for section in sections:
            key = (section["prefix"], section["number"])
            row = [section["crn"]]
            for faculty_name in faculty_names:
                preference = preference_maps[faculty_name].get(key, "X")
                row.append(_normalize_preference_for_csv(preference))
            writer.writerow(row)

    return output_path


def _load_workbook(workbook_path: Path):
    from openpyxl import load_workbook

    return load_workbook(workbook_path, keep_vba=workbook_path.suffix.lower() == ".xlsm")


def _prepare_sheet(workbook, title: str):
    sheet = workbook[title] if title in workbook.sheetnames else workbook.create_sheet(title)
    for merged_range in list(sheet.merged_cells.ranges):
        sheet.unmerge_cells(str(merged_range))
    if sheet.max_row:
        sheet.delete_rows(1, sheet.max_row)
    return sheet


def _style_title_and_headers(sheet, headers: list[str], title: str) -> None:
    from openpyxl.styles import Alignment
    from openpyxl.styles import Font
    from openpyxl.styles import PatternFill

    title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)

    sheet.cell(row=1, column=1).value = title
    sheet.cell(row=1, column=1).fill = title_fill
    sheet.cell(row=1, column=1).font = Font(color="FFFFFF", bold=True, size=14)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))

    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=column_index)
        cell.value = header
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")


def _populate_sections_tab(workbook_path: Path, raw_sections: list[dict[str, str]]) -> None:
    workbook = _load_workbook(workbook_path)
    headers = RAW_SECTION_HEADERS
    sheet = _prepare_sheet(workbook, "Sections")
    _style_title_and_headers(sheet, headers, "Sections")

    for row_index, row_data in enumerate(raw_sections, start=3):
        for column_index, header in enumerate(headers, start=1):
            sheet.cell(row=row_index, column=column_index).value = row_data.get(header, "")

    widths = {
        "A": 10,
        "B": 8,
        "C": 8,
        "D": 8,
        "E": 8,
        "F": 36,
        "G": 12,
        "H": 10,
        "I": 10,
        "J": 14,
        "K": 24,
        "L": 28,
        "M": 16,
        "N": 18,
        "O": 16,
    }
    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = width

    workbook.save(workbook_path)


def _format_time(raw_time: str) -> tuple[str, str]:
    if "-" not in raw_time:
        return raw_time, ""
    start_time, end_time = [value.strip() for value in raw_time.split("-", maxsplit=1)]
    if len(start_time) == 4:
        start_time = f"{start_time[:2]}:{start_time[2:]}"
    if len(end_time) == 4:
        end_time = f"{end_time[:2]}:{end_time[2:]}"
    return start_time, end_time


def _populate_time_tab(workbook_path: Path, raw_sections: list[dict[str, str]]) -> None:
    workbook = _load_workbook(workbook_path)
    headers = ["CRN", "Sub", "Days", "Start Time", "End Time", "Room"]
    sheet = _prepare_sheet(workbook, "Time")
    _style_title_and_headers(sheet, headers, "Time")

    for row_index, row_data in enumerate(raw_sections, start=3):
        start_time, end_time = _format_time(row_data.get("Time", ""))
        row_values = [
            row_data.get("CRN", ""),
            row_data.get("Sub", ""),
            row_data.get("Days", ""),
            start_time,
            end_time,
            row_data.get("Room", ""),
        ]
        for column_index, value in enumerate(row_values, start=1):
            sheet.cell(row=row_index, column=column_index).value = value

    widths = {"A": 10, "B": 8, "C": 10, "D": 12, "E": 12, "F": 24}
    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = width

    workbook.save(workbook_path)


def _populate_preferences_tab(workbook_path: Path, preferences_csv_path: Path) -> None:
    from openpyxl.styles import Alignment
    from openpyxl.styles import Font
    from openpyxl.styles import PatternFill

    workbook = _load_workbook(workbook_path)
    sheet = _prepare_sheet(workbook, "Preferences")

    with preferences_csv_path.open(newline="", encoding="utf-8") as csv_file:
        rows = list(csv.reader(csv_file))

    if not rows:
        rows = [["section_id"]]

    title_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)

    sheet.cell(row=1, column=1).value = "Preferences"
    sheet.cell(row=1, column=1).fill = title_fill
    sheet.cell(row=1, column=1).font = Font(color="FFFFFF", bold=True, size=14)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(rows[0])))

    for row_index, row in enumerate(rows, start=2):
        for column_index, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.value = value
            cell.alignment = Alignment(horizontal="center")
            if row_index == 2:
                cell.fill = header_fill
                cell.font = white_font

    sheet.column_dimensions["A"].width = 14
    for column_index in range(2, len(rows[0]) + 1):
        column_letter = sheet.cell(row=2, column=column_index).column_letter
        sheet.column_dimensions[column_letter].width = 18

    workbook.save(workbook_path)


def build_dean_download_artifacts(term: str | None = None) -> DeanDownloadArtifacts:
    reference_workbook_path = _get_reference_workbook_path()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    sections_csv_path = EXPORT_DIR / f"sections_tab_{timestamp}.csv"
    preferences_csv_path = EXPORT_DIR / f"survey_preferences_{timestamp}.csv"
    workbook_path = EXPORT_DIR / f"faculty_assignment_workbook_{timestamp}{reference_workbook_path.suffix.lower()}"

    sections_csv_path = refresh_bim_sections_csv(sections_csv_path, term=term)
    build_preferences_csv(preferences_csv_path, sections_csv_path)
    shutil.copy2(reference_workbook_path, workbook_path)

    raw_sections = load_raw_sections_tab_data(sections_csv_path)
    _populate_sections_tab(workbook_path, raw_sections)
    _populate_time_tab(workbook_path, raw_sections)
    _populate_preferences_tab(workbook_path, preferences_csv_path)

    return DeanDownloadArtifacts(
        sections_csv_path=sections_csv_path,
        preferences_csv_path=preferences_csv_path,
        workbook_path=workbook_path,
    )
