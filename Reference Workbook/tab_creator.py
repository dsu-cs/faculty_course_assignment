import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import csv

TAB_COLORS = {
    "Sections": {
        "title": "1F4E78",  # Navy
        "header": "5B9BD5", # Steel Blue
        "row_band": "E9EBF5", # Light Grey
        "accent_col": "DCE6F1" # CRN Blue
    },
    "Time": {
        "title": "C65911",    # Deep Burnt Orange (Title)
        "header": "F4B084",   # Soft Muted Orange (Headers)
        "row_band": "F2F2F2", # Neutral Light Grey
        "accent_col": "FFF7ED" # Very Light Peach (CRN Column)
    }
}


# --- 1. DATA LOADING FUNCTION ---
def load_csv_data(csv_path):
    with open(csv_path, mode='r', encoding='utf-8') as f:
        return list(csv.DictReader(f))

def create_tabs(file_path, tab_names):
    workbook = openpyxl.load_workbook(file_path)
    for name in tab_names:
        if name not in workbook.sheetnames:
            workbook.create_sheet(title=name)
            print(f"Created new tab: {name}")
    workbook.save(file_path)


def apply_tab_styles(sheet, excel_headers, total_rows, tab_name):
    # Get the colors for the current tab (default to Sections if not found)
    colors = TAB_COLORS.get(tab_name, TAB_COLORS["Sections"])

    # --- 1. TITLE STYLING (Row 1) ---
    title_fill = PatternFill(start_color=colors["title"], end_color=colors["title"], fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    
    # Fill row 1
    for col in range(1, len(excel_headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = title_fill
        if col == 1:
            cell.font = title_font
            cell.alignment = Alignment(horizontal="left")

    # --- 2. HEADER STYLING (Row 2) ---
    header_fill = PatternFill(start_color=colors["header"], end_color=colors["header"], fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # White side borders for headers
    white_side = Side(style='thin', color="FFFFFF")
    header_border = Border(left=white_side, right=white_side, top=white_side, bottom=white_side)

    for col in range(1, len(excel_headers) + 1):
        header_cell = sheet.cell(row=2, column=col)
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = Alignment(horizontal="center")
        header_cell.border = header_border

    # --- 3. DATA AREA STYLING (Row 3+) ---
    band_fill = PatternFill(start_color=colors["row_band"], end_color=colors["row_band"], fill_type="solid")
    accent_fill = PatternFill(start_color=colors["accent_col"], end_color=colors["accent_col"], fill_type="solid")

    for row_idx in range(3, total_rows + 1):
        for col_idx in range(1, len(excel_headers) + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # 1. Accent Column (CRN)
            if col_idx == 1:
                cell.fill = accent_fill
                cell.font = Font(bold=True)
            
            # 2. Banded Rows
            elif row_idx % 2 == 0:
                cell.fill = band_fill


def populate_sections_tab(file_path, data):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook["Sections"]

    #Title
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = "Section"
    try:
        sheet.merge_cells("A1:J1")
    except:
        pass


    # 1. Define the exact headers for the Excel sheet
    # Added "Waitlist" and "Faculty" to the end
    excel_headers = [
        "CRN", "Sub", "Num", "Seq", "Crd", "Desc", 
        "Current Seats", "Max Seats", "Waitlist", "Faculty"
    ]
    
    # Write headers to Row 2
    for col, header in enumerate(excel_headers, start=1):
        cell = sheet.cell(row=2, column=col)
        cell.value = header

        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # 2. Map the column positions
    # These are the columns we pull directly from CSV keys (including Waitlist and Faculty)
    targets = ["CRN", "Sub", "Num", "Seq", "Crd", "Desc", "Waitlist", "Faculty"]
    col_map = {name: excel_headers.index(name) + 1 for name in targets}
    
    # Indices for our custom split columns
    current_seats_col = excel_headers.index("Current Seats") + 1
    max_seats_col = excel_headers.index("Max Seats") + 1

    # 3. Add Data starting at Row 3
    for row_idx, row_data in enumerate(data, start=3):
        # Fill standard columns (including the new ones)
        for field in targets:
            col_index = col_map[field]
            cell = sheet.cell(row=row_idx, column=col_index)
            cell.value = row_data.get(field, "")
            
            # We center everything UNLESS it is "Desc" or "Faculty"
            if field not in ["Desc", "Faculty"]:
                cell.alignment = Alignment(horizontal="center")
        
        # Split and fill Seats columns
        seats_raw = row_data.get("Seats", "")
        if "/" in seats_raw:
            current_val, max_val = seats_raw.split("/")
            
            # Helper to fill and center the split cells
            for col, val in [(current_seats_col, current_val), (max_seats_col, max_val)]:
                seat_cell = sheet.cell(row=row_idx, column=col)
                seat_cell.value = val
                seat_cell.alignment = Alignment(horizontal="center")

    # 4. Fixed Widths Logic
    # Set the baseline (7) for everything
    for col_index in range(1, len(excel_headers) + 1):
        sheet.column_dimensions[get_column_letter(col_index)].width = 7

    # Updated overrides: Faculty needs to be much wider (approx 25-30)
    overrides = {
        "Desc": 38, 
        "Seq": 5, 
        "Current Seats": 12, 
        "Max Seats": 10,
        "Faculty": 25  # Added this for readability
    }
    
    for name, width in overrides.items():
        if name in excel_headers:
            idx = excel_headers.index(name) + 1
            sheet.column_dimensions[get_column_letter(idx)].width = width

    total_rows = 2 + len(data)        

    apply_tab_styles(sheet, excel_headers, total_rows, "Sections")

    workbook.save(file_path)

def populate_time_tab(file_path, data):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook["Time"]

    blue_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    #Title
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = "Time"
    try:
        sheet.merge_cells("A1:E1")
    except:
        pass

    title_cell.font = Font(bold=True, size=14)
    title_cell.fill = blue_fill

    excel_headers = [
        "CRN", "Sub", "Days", "Start Time", "End Time", "Room"
    ]
    # Write headers to Row 2
    for col, header in enumerate(excel_headers, start=1):
        cell = sheet.cell(row=2, column=col)
        cell.value = header

        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # 2. Map the column positions
    # These are the columns we pull directly from CSV keys (including Waitlist and Faculty)
    targets = ["CRN", "Sub", "Days", "Room"]
    col_map = {name: excel_headers.index(name) + 1 for name in targets}

    # Specific indices for the split time columns
    start_time_col = excel_headers.index("Start Time") + 1
    end_time_col = excel_headers.index("End Time") + 1


    # 3. Add Data starting at Row 3
    for row_idx, row_data in enumerate(data, start=3):
        # Fill standard columns (CRN, Sub, Days, Room)
        for field in targets:
            col_index = col_map[field]
            sheet.cell(row=row_idx, column=col_index).value = row_data.get(field, "")
        
        # --- SPLIT TIME LOGIC ---
        time_raw = row_data.get("Time", "") # Get "1100-1150"
        if "-" in time_raw:
            start_t, end_t = time_raw.split("-")

            def format_time(t):
                t = t.strip()
                if len(t) == 4:
                    return f"{t[:2]}:{t[2:]}"
                return t
            
            sheet.cell(row=row_idx, column=start_time_col).value = format_time(start_t)
            sheet.cell(row=row_idx, column=end_time_col).value = format_time(end_t)
        else:
            sheet.cell(row=row_idx, column=start_time_col).value = time_raw

    # 4. Optional: Set a fixed width so it's readable
    for col_index in range(1, len(excel_headers) + 1):
        sheet.column_dimensions[get_column_letter(col_index)].width = 9


    # Updated overrides: Faculty needs to be much wider (approx 25-30)
    overrides = {
        "CRN": 7,
        "Sub": 7,
        "Room": 30 
    }
    
    for name, width in overrides.items():
        if name in excel_headers:
            idx = excel_headers.index(name) + 1
            sheet.column_dimensions[get_column_letter(idx)].width = width

    total_rows = 2 + len(data)        

    apply_tab_styles(sheet, excel_headers, total_rows, "Time")

    workbook.save(file_path)


def populate_assignment_tab(file_path, data):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook["Assignment"]

    blue_fill = PatternFill(start_color="44b84f", end_color="44b84f", fill_type="solid")

    #Title
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = "Assignment"
    try:
        sheet.merge_cells("A1:E1")
    except:
        pass

    title_cell.font = Font(bold=True, size=14)
    title_cell.fill = blue_fill

    #Headers:  CRN ... ALL TEACHERS NAMES
    workbook.save(file_path)

if __name__ == "__main__":
    csv_file = "sections_tab.csv"
    excel_file = "workbook.xlsx"
    all_tabs = ["Sections", "Time", "Assignment"]

    sections_data = load_csv_data(csv_file)

    create_tabs(excel_file, all_tabs)

    populate_sections_tab(excel_file, sections_data)
    populate_time_tab(excel_file, sections_data)
    populate_assignment_tab(excel_file, sections_data)